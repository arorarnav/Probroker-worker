"""
Builds the deliverable .xlsx: a Listings sheet (same format as the manual
reports we built earlier) plus a new Matches sheet -- the adjacency pairs
that are the actual product differentiator.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
MATCH_HEADER_FILL = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_NORMAL = Font(name="Arial", size=10)


def _style_header_row(ws, headers, fill):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def build_report(rows: list[dict], matches: list[dict], out_path: str,
                  active_days: int = 75, aging_days: int = 150):
    wb = openpyxl.Workbook()

    # ---------- Sheet 1: Listings ----------
    ws = wb.active
    ws.title = "Listings"
    headers = ["Date Posted", "Times Posted", "Posted By", "Type", "Location", "Notes",
               "Size", "Price", "Contact", "Days Since Posted", "Status"]
    _style_header_row(ws, headers, HEADER_FILL)

    ws["M1"] = "Active if age <="
    ws["N1"] = active_days
    ws["M2"] = "Aging if age <="
    ws["N2"] = aging_days

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.get("date")).font = FONT_NORMAL
        ws.cell(row=i, column=2, value=r.get("times_posted", 1)).font = FONT_NORMAL
        ws.cell(row=i, column=3, value=r.get("poster")).font = FONT_NORMAL
        ws.cell(row=i, column=4, value=r.get("listing_type")).font = FONT_NORMAL
        ws.cell(row=i, column=5, value=r.get("location")).font = FONT_NORMAL
        ws.cell(row=i, column=6, value=r.get("notes")).font = FONT_NORMAL
        ws.cell(row=i, column=7, value=r.get("size")).font = FONT_NORMAL
        ws.cell(row=i, column=8, value=r.get("price")).font = FONT_NORMAL
        ws.cell(row=i, column=9, value=r.get("contact")).font = FONT_NORMAL
        ws.cell(row=i, column=10, value=f"=TODAY()-DATEVALUE(A{i})").font = FONT_NORMAL
        ws.cell(row=i, column=11, value=(
            f'=IF(J{i}<=$N$1,"Active",IF(J{i}<=$N$2,"Aging","Stale"))'
        )).font = FONT_NORMAL
        for col in range(1, 12):
            ws.cell(row=i, column=col).border = BORDER

    last_row = len(rows) + 1
    widths = {1: 12, 2: 12, 3: 22, 4: 10, 5: 34, 6: 34, 7: 20, 8: 20, 9: 18, 10: 10, 11: 10}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.add_table(Table(displayName="Listings", ref=f"A1:K{last_row}",
                        tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
    for i in range(2, last_row + 1):
        ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"

    # ---------- Sheet 2: Matches ----------
    ws2 = wb.create_sheet("Matches")
    match_headers = ["Score", "Match Summary", "Demand Date", "Demand Poster", "Demand Location",
                      "Demand Size", "Demand Contact", "Demand Notes", "Supply Date", "Supply Poster",
                      "Supply Location", "Supply Size", "Supply Contact", "Supply Notes"]
    _style_header_row(ws2, match_headers, MATCH_HEADER_FILL)

    for i, m in enumerate(matches, start=2):
        d, s = m["demand"], m["supply"]
        ws2.cell(row=i, column=1, value=m["score"]).font = FONT_NORMAL
        ws2.cell(row=i, column=2, value=m["summary"]).font = FONT_NORMAL
        ws2.cell(row=i, column=3, value=d.get("date")).font = FONT_NORMAL
        ws2.cell(row=i, column=4, value=d.get("poster")).font = FONT_NORMAL
        ws2.cell(row=i, column=5, value=d.get("location")).font = FONT_NORMAL
        ws2.cell(row=i, column=6, value=d.get("size")).font = FONT_NORMAL
        ws2.cell(row=i, column=7, value=d.get("contact")).font = FONT_NORMAL
        ws2.cell(row=i, column=8, value=d.get("notes")).font = FONT_NORMAL
        ws2.cell(row=i, column=9, value=s.get("date")).font = FONT_NORMAL
        ws2.cell(row=i, column=10, value=s.get("poster")).font = FONT_NORMAL
        ws2.cell(row=i, column=11, value=s.get("location")).font = FONT_NORMAL
        ws2.cell(row=i, column=12, value=s.get("size")).font = FONT_NORMAL
        ws2.cell(row=i, column=13, value=s.get("contact")).font = FONT_NORMAL
        ws2.cell(row=i, column=14, value=s.get("notes")).font = FONT_NORMAL
        for col in range(1, 15):
            ws2.cell(row=i, column=col).border = BORDER

    match_widths = {1: 8, 2: 60, 3: 12, 4: 18, 5: 22, 6: 16, 7: 16, 8: 32, 9: 12, 10: 18, 11: 22, 12: 16, 13: 16, 14: 32}
    for col, w in match_widths.items():
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"
    if matches:
        ws2.add_table(Table(displayName="Matches", ref=f"A1:N{len(matches)+1}",
                             tableStyleInfo=TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)))
    for i in range(2, len(matches) + 2):
        ws2.cell(row=i, column=3).number_format = "yyyy-mm-dd"
        ws2.cell(row=i, column=9).number_format = "yyyy-mm-dd"

    wb.save(out_path)
    return out_path
